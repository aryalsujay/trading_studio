#!/usr/bin/env python3
"""
Generate Excel template with dropdown validation for ETF trade imports.
This template includes dropdown validation for the symbol column.
"""

import json
import sys

# For Excel generation, we'll use openpyxl if available
try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

# Fetch symbols from the API
import urllib.request
import urllib.error

try:
    with urllib.request.urlopen('http://localhost:3000/api/symbols') as response:
        symbols_data = json.loads(response.read().decode())
        symbols = sorted([s['symbol'] for s in symbols_data])
except urllib.error.URLError:
    print("WARNING: Could not fetch symbols from API. Using fallback list.")
    symbols = ['NIFTYBEES', 'BANKBEES', 'GOLDBEES', 'ITBEES']  # Fallback

# Create workbook
wb = Workbook()

# Main sheet for data entry
ws = wb.active
ws.title = "Trade Import"

# Header row with styling
headers = ['trade_number', 'symbol', 'entry_date', 'entry_price', 'quantity', 'exit_date', 'exit_price', 'notes']
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Add sample data rows
sample_data = [
    ['1', 'NIFTYBEES', '2026-01-15', '250', '10', '2026-01-20', '260', 'Sample profitable trade'],
    ['', 'BANKBEES', '2026-01-18', '450', '5', '', '', 'LIVE trade - auto-numbered'],
    ['3', 'GOLDBEES', '2026-01-20', '60', '20', '2026-01-25', '58', 'Sample loss trade'],
]

for row_num, row_data in enumerate(sample_data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=value)

# Create a hidden sheet for symbol dropdown list
symbols_sheet = wb.create_sheet("Symbols")
symbols_sheet.sheet_state = 'hidden'

for idx, symbol in enumerate(symbols, 1):
    symbols_sheet.cell(row=idx, column=1, value=symbol)

# Add data validation for symbol column (column B, rows 2 onwards)
# Reference the hidden Symbols sheet
dv = DataValidation(
    type="list",
    formula1=f"=Symbols!$A$1:$A${len(symbols)}",
    allow_blank=False,
    showDropDown=True
)
dv.error = 'Please select a valid ETF symbol from the list'
dv.errorTitle = 'Invalid Symbol'
dv.prompt = 'Choose an ETF symbol from the dropdown'
dv.promptTitle = 'ETF Symbol'

# Apply validation to column B (symbol) for rows 2-1000
ws.add_data_validation(dv)
dv.add(f'B2:B1000')

# Adjust column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 30

# Add instructions sheet
instructions = wb.create_sheet("Instructions", 0)
instructions['A1'] = "ETF Trade Import Template - Instructions"
instructions['A1'].font = Font(size=14, bold=True)

instruction_text = [
    "",
    "How to use this template:",
    "",
    "1. Go to the 'Trade Import' sheet",
    "2. Fill in your trade data row by row",
    "3. For the 'symbol' column, click the dropdown to select from available ETF symbols",
    "4. Leave 'trade_number' blank to auto-number from last database entry",
    "5. For LIVE trades, leave 'exit_date' and 'exit_price' blank",
    "6. Save the file and import via the Trades page",
    "",
    "Column Details:",
    "• trade_number: Optional. Leave blank for auto-numbering",
    "• symbol: REQUIRED. Select from dropdown or type manually",
    "• entry_date: REQUIRED. Format: YYYY-MM-DD (e.g., 2026-01-15)",
    "• entry_price: REQUIRED. Entry price per unit",
    "• quantity: REQUIRED. Number of units",
    "• exit_date: Optional. Format: YYYY-MM-DD",
    "• exit_price: Optional. Exit price per unit",
    "• notes: Optional. Any notes about the trade",
    "",
    f"Available Symbols: {len(symbols)} ETF symbols loaded from database",
    "",
    "Note: New symbols will be automatically added to the database when imported."
]

for row_num, text in enumerate(instruction_text, 2):
    instructions.cell(row=row_num, column=1, value=text)

instructions.column_dimensions['A'].width = 80

# Save the workbook
output_file = '../public/trade-import-template.xlsx'
wb.save(output_file)

print(f"✅ Excel template created successfully: {output_file}")
print(f"📊 {len(symbols)} symbols included in dropdown validation")
